from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.config import get_upload_dir
from app.constants import DOCUMENT_PRIORITIES, DOCUMENT_STATUSES, INCOMING_ACTIONS
from app.models.document import Document
from app.models.document_log import DocumentLog
from app.models.document_type import DocumentType
from app.services.document_service import extract_received_time, search_documents
from app.timezone import format_local_date, format_local_datetime, local_now


BORDER_SIDE = Side(style="thin", color="B7C9D8")
TABLE_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
BODY_FILL = PatternFill("solid", fgColor="FFFFFF")
ALT_BODY_FILL = PatternFill("solid", fgColor="F3F8FC")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")


def _status_label(status: str) -> str:
    return DOCUMENT_STATUSES.get(status, status)


def _priority_label(priority: str) -> str:
    return DOCUMENT_PRIORITIES.get(priority, priority)


def _incoming_action_label(action: str | None) -> str:
    return INCOMING_ACTIONS.get(action or "", action or "")


def _format_plain_date(value: date | None) -> str:
    return value.strftime("%d/%m/%Y") if value else ""


def _latest_logs_by_document(db: Session, document_ids: list[int]) -> dict[int, DocumentLog]:
    if not document_ids:
        return {}
    logs = db.execute(
        select(DocumentLog)
        .options(selectinload(DocumentLog.performer))
        .where(DocumentLog.document_id.in_(document_ids))
        .order_by(DocumentLog.document_id, DocumentLog.performed_at.desc(), DocumentLog.id.desc())
    ).scalars()
    latest_logs: dict[int, DocumentLog] = {}
    for log in logs:
        if log.document_id is not None and log.document_id not in latest_logs:
            latest_logs[log.document_id] = log
    return latest_logs


def _report_document_number(doc: Document) -> str:
    if doc.source_document_id:
        source = doc.source_document
        related_number = ""
        if source:
            related_number = source.document_number or source.document_code or ""
        return f"Văn bản liên quan {related_number.upper()}" if related_number else "none"
    return doc.document_number or "none"


def _document_type_branches(db: Session) -> dict[str, str]:
    rows = db.execute(select(DocumentType.name, DocumentType.branch)).all()
    return {
        name: branch if branch in {"incoming", "outgoing"} else "outgoing"
        for name, branch in rows
    }


def _document_branch(doc: Document, type_branches: dict[str, str]) -> str:
    if doc.document_type and doc.document_type in type_branches:
        return type_branches[doc.document_type]
    if doc.incoming_action:
        return "incoming"
    return "outgoing"


def _last_update_values(doc: Document, latest_logs: dict[int, DocumentLog]) -> tuple[str, str, str]:
    latest_log = latest_logs.get(doc.id)
    last_update_at = latest_log.performed_at if latest_log else doc.updated_at
    last_update_user = (
        latest_log.performer.full_name
        if latest_log and latest_log.performer
        else doc.updater.full_name if doc.updater else ""
    )
    last_note = latest_log.note if latest_log and latest_log.note else ""
    return (
        format_local_datetime(last_update_at) if last_update_at else "",
        last_update_user,
        last_note,
    )


def _permitted_users(doc: Document) -> str:
    return ", ".join(p.user.full_name for p in doc.permissions if p.user)


def _setup_title(ws, title: str, subtitle: str, column_count: int, title_font: Font, normal_font: Font) -> None:
    last_column = get_column_letter(column_count)
    ws.merge_cells(f"A1:{last_column}1")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(f"A2:{last_column}2")
    ws["A2"] = subtitle
    ws["A2"].font = normal_font
    ws["A2"].alignment = Alignment(horizontal="center")


def _style_table_header(
    ws,
    row: int,
    headers: list[str],
    header_font: Font,
    header_fill: PatternFill,
    start_column: int = 1,
) -> None:
    for col, header in enumerate(headers, start=start_column):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = TABLE_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_body_cell(cell, normal_font: Font, row_offset: int, horizontal: str | None = None) -> None:
    cell.font = normal_font
    cell.fill = ALT_BODY_FILL if row_offset % 2 else BODY_FILL
    cell.border = TABLE_BORDER
    cell.alignment = Alignment(horizontal=horizontal, vertical="top", wrap_text=True)


def _write_table(
    ws,
    *,
    start_row: int,
    start_column: int,
    headers: list[str],
    rows: list[list[object]],
    normal_font: Font,
    header_font: Font,
    header_fill: PatternFill,
) -> int:
    _style_table_header(ws, start_row, headers, header_font, header_fill, start_column=start_column)
    for row_offset, row_values in enumerate(rows, start=1):
        for col_offset, value in enumerate(row_values):
            cell = ws.cell(row=start_row + row_offset, column=start_column + col_offset, value=value)
            _style_body_cell(cell, normal_font, row_offset)
    return start_row + max(len(rows), 1)


def _write_section_title(ws, *, row: int, start_column: int, end_column: int, title: str, section_font: Font) -> None:
    for col in range(start_column, end_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = SECTION_FILL
        cell.border = TABLE_BORDER
        cell.font = section_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row, column=start_column, value=title)


def _write_detail_sheet(
    ws,
    *,
    title: str,
    headers: list[str],
    rows: list[list[object]],
    widths: list[int],
    title_font: Font,
    normal_font: Font,
    header_font: Font,
    header_fill: PatternFill,
) -> None:
    ws.sheet_view.showGridLines = False
    _setup_title(
        ws,
        title,
        f"Ngày xuất báo cáo: {local_now():%d/%m/%Y %H:%M}",
        len(headers),
        title_font,
        normal_font,
    )
    header_row = 5
    if rows:
        last_row = _write_table(
            ws,
            start_row=header_row,
            start_column=1,
            headers=headers,
            rows=rows,
            normal_font=normal_font,
            header_font=header_font,
            header_fill=header_fill,
        )
    else:
        _style_table_header(ws, header_row, headers, header_font, header_fill)
        ws.merge_cells(start_row=header_row + 1, start_column=1, end_row=header_row + 1, end_column=len(headers))
        cell = ws.cell(row=header_row + 1, column=1, value="Không có dữ liệu")
        cell.font = normal_font
        cell.fill = ALT_BODY_FILL
        cell.border = TABLE_BORDER
        cell.alignment = Alignment(horizontal="center")
        last_row = header_row + 1

    last_column = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{header_row}:{last_column}{last_row}"
    ws.freeze_panes = f"A{header_row + 1}"
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_summary_sheet(
    ws,
    *,
    documents: list[Document],
    incoming_documents: list[Document],
    outgoing_documents: list[Document],
    title_font: Font,
    normal_font: Font,
    header_font: Font,
    header_fill: PatternFill,
) -> None:
    ws.sheet_view.showGridLines = False
    section_font = Font(name="Times New Roman", size=13, bold=True)
    _setup_title(
        ws,
        "BÁO CÁO QUẢN LÝ VĂN BẢN",
        f"Ngày xuất báo cáo: {local_now():%d/%m/%Y %H:%M}",
        8,
        title_font,
        normal_font,
    )

    summary_rows = [
        ["Tổng số văn bản", len(documents)],
        ["Văn bản nhận về", len(incoming_documents)],
        ["Văn bản đề xuất đi", len(outgoing_documents)],
        ["Tổng file đính kèm", sum(len(doc.files) for doc in documents)],
    ]
    _write_section_title(ws, row=5, start_column=1, end_column=2, title="Tổng quan", section_font=section_font)
    _write_table(
        ws,
        start_row=6,
        start_column=1,
        headers=["Chỉ tiêu", "Số lượng"],
        rows=summary_rows,
        normal_font=normal_font,
        header_font=header_font,
        header_fill=header_fill,
    )

    action_rows = [
        [label, sum(1 for doc in incoming_documents if doc.incoming_action == action)]
        for action, label in INCOMING_ACTIONS.items()
    ]
    _write_section_title(ws, row=5, start_column=4, end_column=5, title="Hướng xử lý văn bản nhận về", section_font=section_font)
    _write_table(
        ws,
        start_row=6,
        start_column=4,
        headers=["Hướng xử lý", "Số lượng"],
        rows=action_rows,
        normal_font=normal_font,
        header_font=header_font,
        header_fill=header_fill,
    )

    status_rows = []
    for status, label in DOCUMENT_STATUSES.items():
        status_rows.append(
            [
                label,
                sum(1 for doc in documents if doc.status == status),
                sum(1 for doc in incoming_documents if doc.status == status),
                sum(1 for doc in outgoing_documents if doc.status == status),
            ]
        )
    status_header_row = 14
    _write_section_title(ws, row=status_header_row - 1, start_column=1, end_column=4, title="Theo trạng thái", section_font=section_font)
    _write_table(
        ws,
        start_row=status_header_row,
        start_column=1,
        headers=["Trạng thái", "Tổng", "Văn bản nhận về", "Văn bản đề xuất đi"],
        rows=status_rows,
        normal_font=normal_font,
        header_font=header_font,
        header_fill=header_fill,
    )

    for col, width in enumerate([32, 14, 4, 32, 14, 4, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def export_documents_excel(db: Session, **filters) -> Path:
    documents = search_documents(db, **filters)
    type_branches = _document_type_branches(db)
    incoming_documents = [doc for doc in documents if _document_branch(doc, type_branches) == "incoming"]
    outgoing_documents = [doc for doc in documents if _document_branch(doc, type_branches) == "outgoing"]
    latest_logs = _latest_logs_by_document(db, [doc.id for doc in documents])
    reports_dir = get_upload_dir() / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    path = reports_dir / f"bao_cao_van_ban_{local_now():%Y%m%d_%H%M%S}.xlsx"

    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Tổng hợp"
    incoming_sheet = wb.create_sheet("Chi tiết văn bản đến")
    outgoing_sheet = wb.create_sheet("Chi tiết văn bản đi")

    title_font = Font(name="Times New Roman", size=16, bold=True)
    normal_font = Font(name="Times New Roman", size=13)
    header_font = Font(name="Times New Roman", size=13, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    incoming_headers = [
        "STT",
        "Số/ký hiệu",
        "Tiêu đề",
        "Người nhận văn bản",
        "Đơn vị gửi đến",
        "Hướng xử lý",
        "Ngày nhận",
        "Giờ nhận",
        "Ngày cập nhật cuối",
        "Người cập nhật cuối",
        "Ghi chú cuối",
        "Số lượng file đính kèm",
        "Người được phân quyền xử lý",
    ]
    incoming_rows: list[list[object]] = []
    for index, doc in enumerate(incoming_documents, start=1):
        last_update_at, last_update_user, last_note = _last_update_values(doc, latest_logs)
        incoming_rows.append(
            [
                index,
                doc.document_number or "none",
                doc.title,
                doc.proposer_name or "",
                doc.sender_department or "",
                _incoming_action_label(doc.incoming_action),
                _format_plain_date(doc.due_date),
                extract_received_time(doc.note),
                last_update_at,
                last_update_user,
                last_note,
                len(doc.files),
                _permitted_users(doc),
            ]
        )

    outgoing_headers = [
        "STT",
        "Văn bản nhận về liên quan",
        "Tiêu đề",
        "Người đề xuất",
        "Phân loại văn bản",
        "Đơn vị nhận văn bản",
        "Mức độ ưu tiên",
        "Trạng thái hiện tại",
        "Ngày hết hạn",
        "Ngày cập nhật cuối",
        "Người cập nhật cuối",
        "Ghi chú cuối",
        "Số lượng file đính kèm",
        "Người được phân quyền xử lý",
    ]
    outgoing_rows: list[list[object]] = []
    for index, doc in enumerate(outgoing_documents, start=1):
        last_update_at, last_update_user, last_note = _last_update_values(doc, latest_logs)
        outgoing_rows.append(
            [
                index,
                _report_document_number(doc),
                doc.title,
                doc.proposer_name or "",
                doc.document_type or "",
                doc.receiver_department or "",
                _priority_label(doc.priority),
                _status_label(doc.status),
                _format_plain_date(doc.due_date),
                last_update_at,
                last_update_user,
                last_note,
                len(doc.files),
                _permitted_users(doc),
            ]
        )

    _write_summary_sheet(
        summary_sheet,
        documents=documents,
        incoming_documents=incoming_documents,
        outgoing_documents=outgoing_documents,
        title_font=title_font,
        normal_font=normal_font,
        header_font=header_font,
        header_fill=header_fill,
    )
    _write_detail_sheet(
        incoming_sheet,
        title="CHI TIẾT VĂN BẢN ĐẾN",
        headers=incoming_headers,
        rows=incoming_rows,
        widths=[8, 22, 36, 22, 24, 24, 16, 14, 20, 22, 42, 18, 36],
        title_font=title_font,
        normal_font=normal_font,
        header_font=header_font,
        header_fill=header_fill,
    )
    _write_detail_sheet(
        outgoing_sheet,
        title="CHI TIẾT VĂN BẢN ĐI",
        headers=outgoing_headers,
        rows=outgoing_rows,
        widths=[8, 28, 36, 22, 22, 24, 18, 24, 16, 20, 22, 42, 18, 36],
        title_font=title_font,
        normal_font=normal_font,
        header_font=header_font,
        header_fill=header_fill,
    )

    wb.save(path)
    return path
